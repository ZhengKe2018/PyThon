import docx
import os
from docx import Document #导入库
from openpyxl import Workbook
from tools import *

workbook = Workbook()
booksheet = workbook.active

PATH = "c:/word/"
files = os.listdir(PATH)
for file in files:
	document = Document(PATH + file) #读入文件
	tables = document.tables #获取文件中的表格集
	table = tables[0]#获取文件中的第一个表格

	list = []

	name = table.cell(0,1).text
	sex = table.cell(0,5).text
	birthday = table.cell(0,10).text
	identity  = table.cell(1,1).text
	ID  = table.cell(1,5).text
	email = table.cell(4,5).text
	phone = table.cell(4,9).text
	addr = table.cell(3,10).text

	edu = table.cell(5,2).text
	work = table.cell(6,2).text

	list.append(name.strip())
	list.append(sex.strip())
	list.append(birthday.strip())
	list.append(identity.strip())
	list.append(ID.strip())
	list.append(email.strip())
	list.append(phone.strip())

	list.append(edu.split("\n")[0].strip())
	list.append(edu.split("\n")[1].strip())
	list.append(edu.split("\n")[2].strip())

	list.append(work.split("\n")[0].strip())
	list.append(work.split("\n")[1].strip())

	booksheet.append(list)

#	print(name + sex + birthday + identity + ID  + email + phone + edu + work)

workbook.save(PATH + 'test.xlsx')
